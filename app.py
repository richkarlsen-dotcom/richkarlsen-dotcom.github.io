"""
Skat Positivliste – Flask backend
Run with: python app.py
Then open: http://localhost:5000
"""

import io
import os
import re
from datetime import datetime
import requests
import openpyxl
from flask import Flask, jsonify, request

app = Flask(__name__)

EXCEL_URL = "https://skat.dk/media/btpf4wfr/februar-2026-abis-liste-2021-2026.xlsx"

_cache = {"rows": None, "headers": None, "isin_col": None}

ISIN_RE = re.compile(r'^[A-Z]{2}[A-Z0-9]{10}$')


def looks_like_isin(val):
    return bool(ISIN_RE.match(str(val).strip().upper()))


def load_data():
    if _cache["rows"] is not None:
        return _cache["headers"], _cache["rows"], _cache["isin_col"]

    print("Downloading Excel from Skat.dk…")
    resp = requests.get(EXCEL_URL, timeout=30)
    resp.raise_for_status()
    raw_bytes = resp.content

    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    all_sheets = wb.sheetnames
    print(f"Sheets: {all_sheets}")

    # Target current year sheet, fallback to highest numeric sheet
    current_year = str(datetime.now().year)
    if current_year in all_sheets:
        sheet_name = current_year
    else:
        year_sheets = [s for s in all_sheets if s.strip().isdigit()]
        sheet_name = max(year_sheets, key=int) if year_sheets else all_sheets[0]

    print(f"Using sheet: '{sheet_name}'")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Find header row — look for a row where a cell contains "ISIN"
    header_idx = 0
    for i, row in enumerate(rows[:30]):
        if any("ISIN" in str(c).upper() for c in row if c is not None):
            header_idx = i
            print(f"Header row found at index {i}")
            break

    headers = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
    print(f"Headers: {headers}")

    data_rows = [
        [str(c).strip() if c is not None else "" for c in row]
        for row in rows[header_idx + 1:]
        if any(c is not None and str(c).strip() != "" for c in row)
    ]

    # Find ISIN column — first try header name, then sniff actual data
    isin_col = next(
        (i for i, h in enumerate(headers) if "ISIN" in h.upper()), None
    )

    if isin_col is None:
        # Sniff: find the column where most values look like ISINs
        print("ISIN not found in headers — sniffing data columns...")
        sample = data_rows[:50]
        col_count = len(headers) if headers else (len(sample[0]) if sample else 0)
        best_col, best_score = 0, 0
        for col in range(col_count):
            score = sum(1 for row in sample if col < len(row) and looks_like_isin(row[col]))
            if score > best_score:
                best_score, best_col = score, col
        isin_col = best_col
        print(f"Sniffed ISIN column as index {isin_col} (score {best_score}/{len(sample)})")
    else:
        print(f"ISIN column found by header name at index {isin_col}")

    _cache["headers"] = headers
    _cache["rows"] = data_rows
    _cache["isin_col"] = isin_col
    print(f"Loaded {len(data_rows)} data rows. ISIN column index: {isin_col}\n")
    return headers, data_rows, isin_col


@app.route("/")
def index():
    with open("index.html", "r", encoding="utf-8") as f:
        return f.read(), 200, {"Content-Type": "text/html"}


@app.route("/api/search")
def search():
    isin = request.args.get("isin", "").strip().upper()
    if not isin:
        return jsonify({"error": "Missing ISIN parameter"}), 400

    try:
        headers, rows, isin_col = load_data()
    except Exception as e:
        return jsonify({"error": f"Could not load data: {e}"}), 502

    matches = [row for row in rows if isin_col < len(row) and row[isin_col].upper() == isin]

    return jsonify({
        "isin": isin,
        "headers": headers,
        "matches": matches,
        "total_rows": len(rows),
    })


@app.route("/api/reload", methods=["POST"])
def reload_cache():
    _cache["rows"] = None
    _cache["headers"] = None
    _cache["isin_col"] = None
    try:
        headers, rows, isin_col = load_data()
        return jsonify({"ok": True, "rows": len(rows), "isin_col": isin_col})
    except Exception as e:
        return jsonify({"error": str(e)}), 502


if __name__ == "__main__":
    try:
        load_data()
    except Exception as e:
        print(f"Warning: could not pre-load data – {e}")
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
